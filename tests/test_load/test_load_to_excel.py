import pandas as pd
from scripts.load.load_to_excel import write_master_excel
import openpyxl


def make_csv(path, data):
    pd.DataFrame(data).to_csv(path, index=False)


def test_write_master_excel_creates_workbook(tmp_path):
    data_dir = tmp_path / "data" / "processed"
    data_dir.mkdir(parents=True, exist_ok=True)

    # 1. Master totals
    master_totals = data_dir / "canonical_question_totals.csv"
    make_csv(
        master_totals,
        {"Canonical Question": ["Q1", "Q2"], "Agree": [1, 2], "Disagree": [0, 1], "Overarching": ["O1", "O2"]},
    )

    # 2. Group summaries
    summary_older = data_dir / "consolidated_questions_older.csv"
    summary_younger = data_dir / "consolidated_questions_younger.csv"
    make_csv(summary_older, {"Canonical Question": ["Q1"], "Agree": [1], "Disagree": [0], "Overarching": ["O1"]})
    make_csv(summary_younger, {"Canonical Question": ["Q2"], "Agree": [2], "Disagree": [1], "Overarching": ["O2"]})

    # 3. Response CSVs (simulate pattern-matching filenames)
    for fname in ["sy23-24_OLDER.csv", "sy23-24_YOUNGER.csv"]:
        make_csv(data_dir / fname, {"Canonical Question": ["Q1"], "Agree": [5], "Disagree": [2]})

    output_master = data_dir / "SF_Master_Summary.xlsx"
    write_master_excel(
        output_master=str(output_master),
        master_totals_path=str(master_totals),
        older_summary_path=str(summary_older),
        younger_summary_path=str(summary_younger),
        responses_dirs={"older": str(data_dir), "younger": str(data_dir)},
    )

    # Check: Excel file created
    assert output_master.exists()

    # Check: sheet names
    wb = openpyxl.load_workbook(output_master)
    sheets = wb.sheetnames
    assert "Master Summary" in sheets
    assert "Older Summary" in sheets
    assert "Younger Summary" in sheets
    # Response sheets: "SY 23-24 Older Responses", "SY 23-24 Younger Responses"
    assert any("Older Responses" in s for s in sheets)
    assert any("Younger Responses" in s for s in sheets)

    # Check: content matches
    ms = wb["Master Summary"]
    # First cell should be "Canonical Question"
    assert ms["A1"].value == "Canonical Question"
    # Older summary tab should have "Agree" column
    older = wb["Older Summary"]
    headers = [cell.value for cell in next(older.iter_rows(min_row=1, max_row=1))]
    assert "Agree" in headers
